import sqlite3
import openpyxl
import os


excel_sheet_name = 'questions.xlsx'
table_name = 'questions'


def createExcel():
    # Check if the file already exists and delete it
    if os.path.exists(excel_sheet_name):
        os.remove(excel_sheet_name)

    # Connect to the SQLite database
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()

    # Execute a query to fetch the column names
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns]

    # Execute a query to fetch the data from the database
    cursor.execute(f"SELECT * FROM {table_name}")
    rows = cursor.fetchall()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column names to the first row of the worksheet
    for col_num, column_name in enumerate(column_names, start=1):
        worksheet.cell(row=1, column=col_num, value=column_name)

    # Write the SQLite data to the Excel worksheet
    for row_num, row_data in enumerate(rows, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook as an Excel file
    workbook.save(excel_sheet_name)

    print('Data exported to excel successfully!')

    # Close the database connection
    cursor.close()
    conn.close()
